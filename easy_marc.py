#!/usr/bin/env python3
"""
EasyMARC - Estrae campi da file UNIMARC ISO 2709 in un foglio Excel.

Legge un file .iso in formato ISO 2709 (UNIMARC/MARC21), applica le
regole di estrazione definite in un file config.json e produce un file
Excel (.xlsx) con una riga per record e una colonna per ogni campo
configurato.

Uso:
    python easy_marc.py <file.iso> [--config config.json] [--output out.xlsx]

Opzioni:
    --config    File JSON con la lista dei campi da estrarre
                (default: config.json nella stessa cartella dello script)
    --output    Percorso del file Excel di output
                (default: stesso nome del .iso ma con estensione .xlsx)
"""

import argparse
import json
import logging
import sys
from datetime import datetime
from pathlib import Path

import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Rimuove sequenze di escape ISO 2022 (es. ESC + 1-3 bytes) e altri
# caratteri di controllo non stampabili che Excel non accetta.
_ESC_RE = re.compile(r"\x1b.")   # ESC + 1 byte (sequenze ISO 2022 corte)
_CTRL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _sanitize(text: str) -> str:
    """Rimuove sequenze di escape MARC e caratteri di controllo non validi in Excel."""
    text = _ESC_RE.sub("", text)
    text = _CTRL_RE.sub("", text)
    return text

# ---------------------------------------------------------------------------
# Costanti ISO 2709
# ---------------------------------------------------------------------------
RS = b"\x1e"   # Record Separator (fine campo)
US = b"\x1f"   # Unit Separator (inizio sottocampo)
GS = b"\x1d"   # Group Separator (fine record)


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

def setup_logging():
    logs_dir = Path(__file__).parent / "logs"
    logs_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = logs_dir / f"easy_marc_{ts}.log"

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    logging.info(f"Log: {log_path}")


# ---------------------------------------------------------------------------
# Parsing ISO 2709
# ---------------------------------------------------------------------------

def _parse_field(raw: bytes) -> dict:
    """
    Analizza un campo dati (tag >= 010).
    Restituisce un dict con:
      - 'ind1', 'ind2': indicatori (stringa 1 char)
      - 'subfields': dict {codice: [lista valori]}
    """
    raw = raw.rstrip(b"\x1e")
    if len(raw) < 2:
        return {"ind1": " ", "ind2": " ", "subfields": {}}

    ind1 = chr(raw[0])
    ind2 = chr(raw[1])
    rest = raw[2:]

    subfields: dict[str, list[str]] = {}
    # I sottocampi iniziano con US (0x1F) seguito da codice (1 byte) + valore
    parts = rest.split(b"\x1f")
    for part in parts:
        if not part:
            continue
        code = chr(part[0])
        value = _sanitize(part[1:].decode("utf-8", errors="replace").strip())
        subfields.setdefault(code, []).append(value)

    return {"ind1": ind1, "ind2": ind2, "subfields": subfields}


def parse_iso2709(path: Path) -> list[dict]:
    """
    Legge un file ISO 2709 e restituisce una lista di record.
    Ogni record è un dict: {tag: [list_of_field_values]}
    - Per campi di controllo (001-009): field_value è una stringa
    - Per campi dati (>=010): field_value è un dict con 'ind1','ind2','subfields'
    """
    data = path.read_bytes()
    records = []
    pos = 0

    while pos < len(data):
        if pos + 5 > len(data):
            break
        try:
            rec_len = int(data[pos : pos + 5])
        except ValueError:
            logging.warning(f"Impossibile leggere la lunghezza del record a pos {pos}, interruzione.")
            break
        if rec_len == 0:
            break

        rec = data[pos : pos + rec_len]
        pos += rec_len

        if len(rec) < 24:
            logging.warning("Record troppo corto, saltato.")
            continue

        leader = rec[:24].decode("latin-1", errors="replace")
        try:
            base_addr = int(leader[12:17])
        except ValueError:
            logging.warning("Base address non valido nel leader, record saltato.")
            continue

        # Directory: da byte 24 fino al primo RS
        dir_end = rec.find(b"\x1e", 24)
        if dir_end < 0:
            logging.warning("Directory non trovata nel record, saltato.")
            continue

        dir_data = rec[24:dir_end]
        n_entries = len(dir_data) // 12

        record: dict[str, list] = {"__leader__": leader}
        for i in range(n_entries):
            entry = dir_data[i * 12 : (i + 1) * 12].decode("latin-1", errors="replace")
            if len(entry) < 12:
                continue
            tag = entry[0:3]
            try:
                length = int(entry[3:7])
                offset = int(entry[7:12])
            except ValueError:
                continue

            field_start = base_addr + offset
            field_raw = rec[field_start : field_start + length]

            # Campi di controllo: nessun indicatore né sottocampi
            if tag < "010":
                value = _sanitize(field_raw.rstrip(b"\x1e").decode("utf-8", errors="replace").strip())
            else:
                value = _parse_field(field_raw)

            record.setdefault(tag, []).append(value)

        records.append(record)

    return records


# ---------------------------------------------------------------------------
# Applicazione config → valori cella
# ---------------------------------------------------------------------------

def _apply_formats(field_value: dict, formats: list[dict]) -> str:
    """
    Prova i formati nell'ordine dato e restituisce il primo valido
    (tutti i sottocampi richiesti presenti nel campo).
    """
    subfields = field_value.get("subfields", {})
    for fmt in formats:
        required = fmt.get("subfields", [])
        # Controlla che tutti i sottocampi richiesti siano presenti
        if all(sf in subfields for sf in required):
            template = fmt.get("format", "")
            # Sostituisce {x} con il valore del sottocampo x.
            # Se il sottocampo è in "join", tutti i valori ripetuti vengono
            # uniti con il separatore indicato; altrimenti si usa solo il primo.
            join_seps = fmt.get("join", {})
            result = template
            for sf in required:
                if sf in join_seps and subfields[sf]:
                    val = join_seps[sf].join(subfields[sf])
                else:
                    val = subfields[sf][0] if subfields[sf] else ""
                result = result.replace("{" + sf + "}", val)
            # Slice opzionale sul risultato finale (es. per estrarre anni da 100$a)
            slc = fmt.get("slice")
            if slc:
                result = result[slc[0]:slc[1]]
            return result
    return ""


def _apply_auto(field_value: dict) -> str:
    """Concatena tutti i sottocampi presenti come $x valore."""
    if isinstance(field_value, str):
        return field_value
    subfields = field_value.get("subfields", {})
    parts = []
    for code, values in subfields.items():
        for v in values:
            parts.append(f"${code} {v}")
    return " ".join(parts)


def apply_column(record: dict, col_spec: dict) -> str:
    """
    Estrae il valore di una colonna per un singolo record.
    Gestisce campi ripetuti unendoli con il separatore configurato.
    """
    # Valore costante: restituisce sempre la stringa fissa
    if "constant" in col_spec:
        return col_spec["constant"]

    # Estrazione dal Leader: legge il carattere alla posizione indicata
    if col_spec.get("source") == "leader":
        leader = record.get("__leader__", "")
        offset = col_spec.get("offset", 0)
        val = leader[offset] if offset < len(leader) else ""
        return col_spec.get("map", {}).get(val, val)

    tag = col_spec["tag"]
    separator = col_spec.get("separator", " | ")
    formats = col_spec.get("formats", None)

    occurrences = record.get(tag, [])
    if not occurrences:
        return ""

    flt = col_spec.get("filter")  # es. {"subfield": "3", "value": "ABR0ME"}

    results = []
    for field_value in occurrences:
        # Filtra per valore di sottocampo se richiesto
        if flt and isinstance(field_value, dict):
            sf_code = flt.get("subfield", "")
            sf_val  = flt.get("value", "")
            if sf_val not in field_value.get("subfields", {}).get(sf_code, []):
                continue
        if isinstance(field_value, str):
            # Campo di controllo
            results.append(field_value)
        elif formats:
            val = _apply_formats(field_value, formats)
            if val:
                results.append(val)
        else:
            val = _apply_auto(field_value)
            if val:
                results.append(val)

    return separator.join(results)


# ---------------------------------------------------------------------------
# Scrittura Excel
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
DATA_FONT = Font(name="Calibri", size=10)


def write_excel(records: list[dict], columns: list[dict], output_path: Path):
    """Scrive il file Excel con header e dati."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UNIMARC"

    # Header
    for col_idx, col_spec in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_spec.get("label", col_spec.get("tag", "")))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    ws.row_dimensions[1].height = 18

    # Dati
    for row_idx, record in enumerate(records, start=2):
        for col_idx, col_spec in enumerate(columns, start=1):
            value = apply_column(record, col_spec)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    # Auto-larghezza colonne (stima)
    for col_idx, col_spec in enumerate(columns, start=1):
        label = col_spec.get("label", col_spec.get("tag", ""))
        max_len = len(label)
        for row_idx in range(2, len(records) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value or ""
            max_len = max(max_len, len(str(val)))
        # Limita a 60 caratteri
        adjusted = min(max_len + 2, 62)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI e main
# ---------------------------------------------------------------------------

def load_config(config_path: Path) -> list[dict]:
    """Legge il file JSON e restituisce la lista delle colonne."""
    with open(config_path, encoding="utf-8") as f:
        data = json.load(f)
    columns = data.get("columns", [])
    if not columns:
        raise ValueError("Il file config.json non contiene nessuna colonna ('columns' vuoto o assente).")
    return columns


def main():
    setup_logging()

    parser = argparse.ArgumentParser(
        description="EasyMARC — Estrae campi UNIMARC ISO 2709 in Excel"
    )
    parser.add_argument("iso_file", help="File UNIMARC in formato ISO 2709 (.iso)")
    parser.add_argument(
        "--config",
        default=None,
        help="File JSON di configurazione (default: config.json nella cartella dello script)",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Percorso del file Excel di output (default: stesso nome del .iso con estensione .xlsx)",
    )
    args = parser.parse_args()

    iso_path = Path(args.iso_file)
    if not iso_path.exists():
        logging.error(f"File non trovato: {iso_path}")
        sys.exit(1)

    # Config
    if args.config:
        config_path = Path(args.config)
    else:
        config_path = Path(__file__).parent / "config.json"

    if not config_path.exists():
        logging.error(f"File di configurazione non trovato: {config_path}")
        sys.exit(1)

    try:
        columns = load_config(config_path)
    except (json.JSONDecodeError, ValueError) as e:
        logging.error(f"Errore nel file di configurazione: {e}")
        sys.exit(1)

    logging.info(f"Configurazione: {len(columns)} colonne da {config_path.name}")

    # Output path
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = iso_path.with_suffix(".xlsx")

    # Parsing
    logging.info(f"Leggo: {iso_path}")
    try:
        records = parse_iso2709(iso_path)
    except Exception as e:
        logging.error(f"Errore nel parsing del file ISO: {e}")
        sys.exit(1)

    logging.info(f"Record trovati: {len(records)}")

    # Scrittura Excel
    logging.info(f"Scrivo: {output_path}")
    try:
        write_excel(records, columns, output_path)
    except Exception as e:
        logging.error(f"Errore nella scrittura Excel: {e}")
        sys.exit(1)

    logging.info(f"✓ {len(records)} record elaborati → {output_path}")


if __name__ == "__main__":
    main()
